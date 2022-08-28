import openpyxl, os, zipfile
import numpy as np
import shutil



directory = 'SHANTI AUTO'

working_directory = 'C:/Users/DELL/PycharmProjects/pythonProject/'
os.chdir(working_directory)

#print(os.getcwd())
for root, dirs, files in os.walk(directory):
    print(root, dirs, files)
    for FY in dirs:
        print(FY)
        path = os.path.join(working_directory + directory, FY)
        os.chdir(path)
        #print(os.getcwd())
        new_directory = 'excel_files'
        shutil.rmtree(new_directory, ignore_errors=True)              #delete folder if already exists
        os.mkdir(new_directory)
        os.chdir(path)
        files_list = os.listdir()
        #print(files_list)
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        n_files = len(files_list)
        #print(n_files)
        for file_name in files_list:
            #print(file_name)
            # opening the zip file in READ mode
            if zipfile.is_zipfile(file_name):  # if it is a zipfile, extract it
                with zipfile.ZipFile(file_name, 'r') as zip:
                    # printing all the contents of the zip file
                    #zip.printdir()
                    # extracting all the files
                    #print('Extracting all the files now...')
                    os.chdir(new_directory)
                    zip.extractall()
                    os.chdir(path)
                    #print('Done!')
        #print(os.path.join(os.getcwd(), root))
        k1 = 1
        os.chdir(new_directory)
        #print(os.getcwd())
        files_list = os.listdir()
        n_files = len(files_list)
        for i in range(n_files):
            #print(files_list[i])
            wb = openpyxl.load_workbook(files_list[i])
            ws = wb['B2B']
            #print(ws.max_column, ws.max_row)
            row_start = 7
            if i == 0:
                row_start = 5
            for k in range(row_start, ws.max_row+1):
                for l in range(1, ws.max_column+1):
                    new_worksheet.cell(row = k1 + k - row_start, column = l).value = ws.cell(k ,l).value
            k1 = k1 + k
        #print(new_worksheet.max_row)
        print("deleting blank rows:" + path)
        n_deleted_rows = 0
        for rows in range(1, new_worksheet.max_row + 1):
            if new_worksheet.cell(rows-n_deleted_rows, 3).value is None:
                new_worksheet.delete_rows(rows-n_deleted_rows, 1)
                n_deleted_rows = n_deleted_rows + 1
            if new_worksheet.cell(rows-n_deleted_rows, 9).value == '-':
                new_worksheet.delete_rows(rows - n_deleted_rows, 1)
                n_deleted_rows = n_deleted_rows + 1
        summary = new_workbook.create_sheet("Summary", 0)
        summary.cell(1, 1).value = "Total taxable value"
        summary.cell(1, 2).value = "Total Tax Value"
        summary.cell(1, 3).value = "Avg Tax rate"
        summary.cell(2, 1).value = '=SUM(Sheet!J:J)'
        summary.cell(2, 2).value = '=SUM(Sheet!K:M)'
        summary.cell(2, 3).value = '= B2/A2'
        new_workbook.save('consolidated' + FY + '.xlsx')
        #print(ws.cell(j,i).value)
        #cell[i-1][j-1] = ws.cell(i,j).value

